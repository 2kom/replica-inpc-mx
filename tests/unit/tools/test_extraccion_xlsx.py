from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest
from canasta_inpc.esquema import LAYOUTS_XLSX, VersionCanasta
from canasta_inpc.extraccion_xlsx import extraer_xlsx
from canasta_inpc.utilidades import guardar_csv

# fixtures reales de INEGI, no versionadas (data/ esta en .gitignore) -- si
# faltan, estos tests fallan con FileNotFoundError, no se saltan solos
_DIR_XLSX = Path(__file__).parent.parent.parent.parent / "data" / "tests" / "xlsx"
_DIR_REFERENCIA = Path(__file__).parent.parent.parent.parent / "data" / "tests" / "ponderadores"

_VERSIONES: tuple[VersionCanasta, ...] = (2010, 2013, 2018, 2024)
_VERSIONES_CON_ENCADENAMIENTO: tuple[VersionCanasta, ...] = (2013, 2024)

_CONTEO_ESPERADO: dict[VersionCanasta, int] = {2010: 283, 2013: 283, 2018: 299, 2024: 292}

_COLUMNAS_COMUNES = frozenset(
    {
        "generico",
        "ponderador",
        "COG",
        "CCIF division",
        "inflacion componente",
        "inflacion subcomponente",
        "inflacion agrupacion",
        "canasta basica",
    }
)
_COLUMNAS_ESPERADAS: dict[VersionCanasta, frozenset[str]] = {
    2010: _COLUMNAS_COMUNES,
    2013: _COLUMNAS_COMUNES | {"encadenamiento"},
    2018: _COLUMNAS_COMUNES,
    2024: _COLUMNAS_COMUNES | {"encadenamiento", "canasta consumo minimo"},
}

_SALTAR = {"indice general", "total", "suma", "factor de encadenamiento"}

# genericos donde CCIF division del xlsx 2018 ("ropa y calzado") diverge del
# nombre CCIF oficial ("prendas de vestir y calzado", solo sale del pdf) --
# snapshot exacto, verificado contra data/tests/ponderadores/ponderadores_2018.csv
_CCIF_2018_DIVERGENTES = frozenset(
    {
        "blusas y playeras para mujer",
        "calcetas medias y pantimedias",
        "calcetines y calcetas para hombre",
        "calcetines y calcetas para niños",
        "camisas y playeras para hombre",
        "camisas y playeras para niños",
        "camisetas para bebes",
        "otras prendas de vestir para hombre",
        "otras prendas de vestir para mujer",
        "pantalones para hombre",
        "pantalones para mujer",
        "pantalones para niño",
        "ropa de abrigo",
        "ropa interior para hombre",
        "ropa interior para infantes",
        "ropa interior para mujer",
        "ropa para bebes",
        "servicio de lavanderia",
        "servicio de tintoreria",
        "servicios y articulos para el calzado",
        "traje para hombre",
        "uniformes escolares",
        "vestidos y faldas para mujer",
        "vestidos faldas y pantalones para niñas",
        "zapatos de material sintetico",
        "zapatos para hombre",
        "zapatos para mujer",
        "zapatos para niños y niñas",
        "zapatos tenis",
    }
)


# -- helpers ------------------------------------------------------------


def _extraer(version: VersionCanasta) -> pd.DataFrame:
    return extraer_xlsx(_DIR_XLSX / f"{version}.xlsx", version)


def _referencia(version: VersionCanasta) -> pd.DataFrame:
    ruta = _DIR_REFERENCIA / f"ponderadores_{version}.csv"
    return pd.read_csv(ruta, dtype=str, keep_default_na=False)


def _cruce_con_referencia(version: VersionCanasta) -> pd.DataFrame:
    """Cruza el resultado de extraer_xlsx con el csv de referencia por generico."""
    return _extraer(version).merge(_referencia(version), on="generico", suffixes=("_x", "_r"))


def _traducir_binaria(serie: pd.Series) -> pd.Series:
    """Traduce la convencion vieja de la referencia ("X"/"") a la nueva ("X"/"-")."""
    return serie.apply(lambda v: "X" if v == "X" else "-")


def _filas_con_multiples_marcas(version: VersionCanasta) -> int:
    """Cuenta filas de la hoja COG con 2+ columnas de agrupacion marcadas "X" a la vez."""
    layout = LAYOUTS_XLSX[version]
    wb = openpyxl.load_workbook(_DIR_XLSX / f"{version}.xlsx", data_only=True)
    ws = wb[layout.hoja_cog]
    return sum(
        1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row)
        if sum(1 for col in layout.agrupaciones if row[col - 1].value == "X") >= 2
    )


# -- forma del resultado -------------------------------------------------


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_conteo_de_filas_coincide_con_la_referencia(version: VersionCanasta) -> None:
    assert len(_extraer(version)) == _CONTEO_ESPERADO[version] == len(_referencia(version))


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_columnas_del_resultado_coinciden_con_lo_esperado(version: VersionCanasta) -> None:
    # tambien protege que "_grupo" (columna interna de trabajo) no se filtre
    # al resultado final
    assert set(_extraer(version).columns) == _COLUMNAS_ESPERADAS[version]


# -- generico -------------------------------------------------------------


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_genericos_coinciden_exactamente_con_la_referencia(version: VersionCanasta) -> None:
    assert set(_extraer(version)["generico"]) == set(_referencia(version)["generico"])


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_generico_es_unico_por_version(version: VersionCanasta) -> None:
    assert _extraer(version)["generico"].is_unique


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_generico_nunca_es_encabezado_o_fila_de_totales(version: VersionCanasta) -> None:
    assert not set(_extraer(version)["generico"]) & _SALTAR


# -- ponderador / encadenamiento ------------------------------------------


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_ponderador_coincide_numericamente_con_la_referencia(version: VersionCanasta) -> None:
    # igualdad exacta, no pytest.approx: ambos lados son el mismo valor
    # parseado de texto, sin aritmetica de por medio -- una tolerancia
    # relativa (~1e-6 default) esconderia una deriva real de precision
    m = _cruce_con_referencia(version)
    assert (m["ponderador_x"].astype(float).to_numpy() == m["ponderador_r"].astype(float).to_numpy()).all()


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_ponderador_siempre_positivo(version: VersionCanasta) -> None:
    assert (_extraer(version)["ponderador"].astype(float) > 0).all()


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_suma_de_ponderador_es_100(version: VersionCanasta) -> None:
    # tolerancia explicita y estrecha (no el default relativo ~1e-6 de
    # pytest.approx): la desviacion real por suma flotante de ~283-299
    # terminos es de orden 1e-13, abs=1e-9 deja margen sin esconder un bug real
    assert _extraer(version)["ponderador"].astype(float).sum() == pytest.approx(100, abs=1e-9)


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES_CON_ENCADENAMIENTO)
def test_encadenamiento_coincide_numericamente_con_la_referencia(version: VersionCanasta) -> None:
    # igualdad exacta -- mismo motivo que test_ponderador_coincide_numericamente_con_la_referencia
    m = _cruce_con_referencia(version)
    assert (
        m["encadenamiento_x"].astype(float).to_numpy() == m["encadenamiento_r"].astype(float).to_numpy()
    ).all()


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES_CON_ENCADENAMIENTO)
def test_encadenamiento_siempre_positivo(version: VersionCanasta) -> None:
    assert (_extraer(version)["encadenamiento"].astype(float) > 0).all()


@pytest.mark.requires_data
def test_ponderador_preserva_precision_cruda_del_xml_2018() -> None:
    # celda B16 ("Objeto de gasto", "Alimentos para bebe"): openpyxl parsea
    # 0.03094422504321854, el XML crudo trae "3.0944225043218539E-2" -- mismo
    # double, texto distinto (notacion cientifica vs decimal). El objetivo de
    # _valores_crudos es preservar el segundo, no el primero.
    df = _extraer(2018)
    fila = df.loc[df["generico"] == "alimentos para bebe", "ponderador"]
    assert fila.iloc[0] == "3.0944225043218539E-2"


@pytest.mark.requires_data
def test_encadenamiento_preserva_precision_cruda_del_xml() -> None:
    # mismo mecanismo que ponderador (ver test_ponderador_preserva_precision_cruda_del_xml_2018),
    # verificado aparte porque _valores_crudos podria cambiar a str(cell.value)
    # solo para ponderador sin que ese test lo detecte
    df_2013 = _extraer(2013)
    fila_2013 = df_2013.loc[df_2013["generico"] == "agua embotellada", "encadenamiento"]
    assert fila_2013.iloc[0] == "0.99939577475004904"  # D20, openpyxl da 0.999395774750049

    df_2024 = _extraer(2024)
    fila_2024 = df_2024.loc[
        df_2024["generico"] == "aceites y grasas vegetales comestibles", "encadenamiento"
    ]
    assert fila_2024.iloc[0] == "1.7325362471870001"  # C14, openpyxl da 1.732536247187


# -- COG / inflacion --------------------------------------------------------


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_cog_coincide_con_la_referencia(version: VersionCanasta) -> None:
    m = _cruce_con_referencia(version)
    assert (m["COG_x"] == m["COG_r"]).all()


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
@pytest.mark.parametrize(
    "columna", ["inflacion componente", "inflacion subcomponente", "inflacion agrupacion"]
)
def test_inflacion_coincide_con_la_referencia(version: VersionCanasta, columna: str) -> None:
    m = _cruce_con_referencia(version)
    assert (m[f"{columna}_x"] == m[f"{columna}_r"]).all()


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_inflacion_componente_nunca_vacio(version: VersionCanasta) -> None:
    assert (_extraer(version)["inflacion componente"] != "").all()


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_agrupaciones_del_xlsx_nunca_marcan_dos_categorias_a_la_vez(
    version: VersionCanasta,
) -> None:
    # _clasificar_inflacion toma la primera columna marcada "X" y calla si
    # hay 2+ -- este test protege la fuente cruda (el xlsx real), no el
    # codigo de extraccion, que nunca revelaria la ambiguedad si existiera
    assert _filas_con_multiples_marcas(version) == 0


# -- CCIF division ----------------------------------------------------------


@pytest.mark.requires_data
def test_ccif_division_2010_vacio_por_ausencia_de_hoja_ccif() -> None:
    assert (_extraer(2010)["CCIF division"] == "").all()


@pytest.mark.requires_data
def test_ccif_division_2013_coincide_con_la_referencia() -> None:
    m = _cruce_con_referencia(2013)
    assert (m["CCIF division_x"] == m["CCIF division_r"]).all()


@pytest.mark.requires_data
def test_ccif_division_2018_diverge_de_la_referencia_solo_en_genericos_conocidos() -> None:
    # no es bug de extraccion_xlsx.py: el xlsx trae "ropa y calzado", el
    # nombre CCIF oficial ("prendas de vestir y calzado") solo sale del pdf.
    # Se corrige cuando exista extraer_pdf.py. Snapshot exacto (no solo
    # conteo) para detectar si la divergencia crece o cambia de genericos.
    m = _cruce_con_referencia(2018)
    divergentes = m.loc[m["CCIF division_x"] != m["CCIF division_r"]]
    assert set(divergentes["generico"]) == _CCIF_2018_DIVERGENTES
    assert (divergentes["CCIF division_x"] == "ropa y calzado").all()
    assert (divergentes["CCIF division_r"] == "prendas de vestir y calzado").all()


@pytest.mark.requires_data
def test_ccif_division_2024_coincide_con_la_referencia() -> None:
    m = _cruce_con_referencia(2024)
    assert (m["CCIF division_x"] == m["CCIF division_r"]).all()


@pytest.mark.requires_data
def test_ccif_division_2024_nunca_conserva_prefijo_numerico_del_xlsx() -> None:
    # 2024 es la unica version cuyo xlsx trae el prefijo en la celda CCIF
    # (ej. "01 alimentos..."); extraer_xlsx debe quitarlo siempre -- el
    # prefijo consistente en las 4 versiones lo repone extraer_pdf.py
    assert not _extraer(2024)["CCIF division"].str.match(r"^\d").any()


# -- canasta basica / consumo minimo -----------------------------------------


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_canasta_basica_coincide_con_la_referencia(version: VersionCanasta) -> None:
    # la referencia usa la convencion vieja ("X"/""); esquema.py decidio esta
    # sesion pasar a "X"/"-" -- se traduce antes de comparar
    m = _cruce_con_referencia(version)
    assert (m["canasta basica_x"] == _traducir_binaria(m["canasta basica_r"])).all()


@pytest.mark.requires_data
def test_canasta_consumo_minimo_coincide_con_la_referencia_2024() -> None:
    m = _cruce_con_referencia(2024)
    assert (m["canasta consumo minimo_x"] == _traducir_binaria(m["canasta consumo minimo_r"])).all()


# -- reproducibilidad ---------------------------------------------------


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_extraer_xlsx_es_reproducible(version: VersionCanasta) -> None:
    pd.testing.assert_frame_equal(_extraer(version), _extraer(version))


@pytest.mark.requires_data
@pytest.mark.parametrize("version", _VERSIONES)
def test_csv_escrito_es_byte_identico_entre_corridas(
    version: VersionCanasta, tmp_path: Path
) -> None:
    # el invariante real que le importa al pipeline es el archivo en disco,
    # no solo el DataFrame en memoria -- guardar_csv es el que determina eso
    ruta_a = tmp_path / "a.csv"
    ruta_b = tmp_path / "b.csv"
    guardar_csv(_extraer(version), ruta_a, version)
    guardar_csv(_extraer(version), ruta_b, version)
    assert ruta_a.read_bytes() == ruta_b.read_bytes()
