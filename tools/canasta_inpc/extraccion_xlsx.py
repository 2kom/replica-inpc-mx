import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from canasta_inpc.esquema import LAYOUTS_XLSX, LayoutXlsx, VersionCanasta
from canasta_inpc.utilidades import normalizar_texto, quitar_prefijo_numerico

_SALTAR = {"indice general", "total", "suma", "factor de encadenamiento"}

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def extraer_xlsx(
    ruta: Path,
    version: VersionCanasta,
) -> pd.DataFrame:
    """Extrae genericos, ponderadores y clasificaciones de un xlsx INEGI.

    Ver: tools/uso_generar_canasta.md §Esquema del CSV de salida,
    §Hojas esperadas por version.
    """
    layout = LAYOUTS_XLSX[version]
    wb = openpyxl.load_workbook(ruta, data_only=True)

    df = _leer_hoja(wb[layout.hoja_cog], layout, _valores_crudos(ruta, layout.hoja_cog))

    if layout.hoja_ccif:
        df_ccif = _leer_hoja(wb[layout.hoja_ccif], layout, _valores_crudos(ruta, layout.hoja_ccif))
        mapa_ccif = dict(zip(df_ccif["generico"], df_ccif["_grupo"]))
        df["CCIF division"] = (
            df["generico"].map(mapa_ccif).fillna("").apply(quitar_prefijo_numerico)
        )
    else:
        df["CCIF division"] = ""

    df["COG"] = df["_grupo"].apply(quitar_prefijo_numerico)
    df["generico"] = df["generico"].apply(quitar_prefijo_numerico)

    return df.drop(columns="_grupo")


def _leer_hoja(ws: Worksheet, layout: LayoutXlsx, crudos: dict[str, str]) -> pd.DataFrame:
    """Recorre una hoja y arma una fila por generico valido, con su `_grupo` (categoria)."""
    grupo_actual = ""
    filas: list[dict] = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if not isinstance(_celda(row, layout.col_ponderador), (int, float)):
            continue

        if layout.col_grupo is not None:
            texto_grupo = _texto(_celda(row, layout.col_grupo))
            texto_generico = _texto(_celda(row, layout.col_generico))

            if texto_grupo and not texto_generico:
                if texto_grupo.lower() not in _SALTAR:
                    grupo_actual = normalizar_texto(texto_grupo)
                continue

            if not texto_generico or texto_generico.lower() in _SALTAR:
                continue
            generico = texto_generico
        else:
            texto_generico = _texto(_celda(row, layout.col_generico))
            if not texto_generico or texto_generico.lower() in _SALTAR:
                continue

            tiene_marca = any(_celda(row, col) == "X" for col in layout.agrupaciones)
            if not tiene_marca:
                grupo_actual = normalizar_texto(texto_generico)
                continue
            generico = texto_generico

        comp, subcomp, agrup = _clasificar_inflacion(row, layout)

        fila: dict = {
            "generico": normalizar_texto(generico),
            "ponderador": _celda_cruda(row, layout.col_ponderador, crudos),
            "_grupo": grupo_actual,
            "inflacion componente": comp,
            "inflacion subcomponente": subcomp,
            "inflacion agrupacion": agrup,
            "canasta basica": "X" if _celda(row, layout.col_canasta_basica) == "X" else "-",
        }

        if layout.col_encadenamiento is not None:
            fila["encadenamiento"] = _celda_cruda(row, layout.col_encadenamiento, crudos)

        if layout.col_canasta_consumo_minimo is not None:
            es_ccm = _celda(row, layout.col_canasta_consumo_minimo) == "X"
            fila["canasta consumo minimo"] = "X" if es_ccm else "-"

        filas.append(fila)

    return pd.DataFrame(filas)


def _celda(row: tuple, col: int) -> object:
    """Valor (ya parseado por openpyxl) de la celda en `col` (1-indexado)."""
    return row[col - 1].value


def _celda_cruda(row: tuple, col: int, crudos: dict[str, str]) -> str:
    """Texto crudo del XML para la celda en `col` -- precision exacta, sin pasar por float."""
    return crudos.get(row[col - 1].coordinate, "")


def _texto(valor: object) -> str:
    """Convierte una celda a string, vacio si es None."""
    if valor is None:
        return ""
    return str(valor).strip()


def _clasificar_inflacion(row: tuple, layout: LayoutXlsx) -> tuple[str, str, str]:
    """Busca la columna de agrupacion marcada con "X" y devuelve su clasificacion."""
    for col, clasificacion in layout.agrupaciones.items():
        if _celda(row, col) == "X":
            return clasificacion
    return "", "", ""


def _valores_crudos(ruta: Path, nombre_hoja: str) -> dict[str, str]:
    """Lee el texto crudo (sin parsear a float) de las celdas numericas de una hoja.

    openpyxl parsea a float y no siempre preserva la representacion exacta del
    XML (ej. notacion cientifica se vuelve decimal, ver
    tools/uso_generar_canasta.md §Esquema del CSV de salida). Se usa solo para
    `ponderador`/`encadenamiento`, donde la precision importa.
    """
    with zipfile.ZipFile(ruta) as zf:
        xml = zf.read(_nombre_archivo_hoja(zf, nombre_hoja))

    crudos: dict[str, str] = {}
    for celda in ET.fromstring(xml).iter(f"{{{_NS_MAIN}}}c"):
        tipo = celda.get("t")
        if tipo is not None and tipo != "n":
            continue
        ref = celda.get("r")
        valor = celda.find(f"{{{_NS_MAIN}}}v")
        if ref is not None and valor is not None and valor.text is not None:
            crudos[ref] = valor.text
    return crudos


def _nombre_archivo_hoja(zf: zipfile.ZipFile, nombre_hoja: str) -> str:
    """Resuelve un nombre de hoja (ej. "CCIF") al `sheetN.xml` correspondiente."""
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rid = next(
        s.get(f"{{{_NS_REL}}}id")
        for s in workbook.iter(f"{{{_NS_MAIN}}}sheet")
        if s.get("name") == nombre_hoja
    )
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    destino = next(
        r.get("Target") for r in rels.iter(f"{{{_NS_PKG_REL}}}Relationship") if r.get("Id") == rid
    )
    return f"xl/{destino}"
