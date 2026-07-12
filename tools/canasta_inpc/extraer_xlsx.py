from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from canasta_inpc.config import LAYOUTS_XLSX, LayoutXlsx, columnas_xlsx

_SKIP = {"indice general", "total", "suma", "factor de encadenamiento"}


def extraer_xlsx(ruta: Path, version: int) -> pd.DataFrame:
    """Extrae genéricos, ponderadores y clasificaciones de un xlsx INEGI.

    Ver: tools/uso_generar_canasta.md §Versiones soportadas
    """
    layout = LAYOUTS_XLSX[version]
    wb = openpyxl.load_workbook(ruta, data_only=True)

    df = _leer_hoja(wb[layout.hoja_cog], layout)
    df.rename(columns={"_grupo": "COG"}, inplace=True)

    if layout.hoja_ccif:
        df_ccif = _leer_hoja(wb[layout.hoja_ccif], layout)
        mapa_ccif = dict(zip(df_ccif["generico"], df_ccif["_grupo"]))
        df["CCIF division"] = df["generico"].map(mapa_ccif).fillna("")

    cols = columnas_xlsx(version)
    for col in cols:
        if col not in df.columns:
            df[col] = ""

    return df[cols]


def _leer_hoja(ws: Worksheet, layout: LayoutXlsx) -> pd.DataFrame:
    """Recorre una hoja del xlsx y arma una fila por genérico válido."""
    grupo_actual = ""
    filas: list[dict] = []
    cols_sep = layout.col_concepto_agg != layout.col_concepto_gen

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ponderador = _celda(row, layout.col_ponderador)
        if not isinstance(ponderador, (int, float)):
            continue

        if cols_sep:
            concepto_agg = _str(_celda(row, layout.col_concepto_agg))
            concepto_gen = _str(_celda(row, layout.col_concepto_gen))

            if concepto_agg and not concepto_gen:
                if not _es_skip(concepto_agg):
                    grupo_actual = concepto_agg
                continue

            if not concepto_gen or _es_skip(concepto_gen):
                continue
            concepto = concepto_gen
        else:
            concepto = _str(_celda(row, layout.col_concepto_agg))
            if not concepto or _es_skip(concepto):
                continue

            tiene_x = any(_celda(row, c) == "X" for c in layout.agrupaciones)
            if not tiene_x:
                grupo_actual = concepto
                continue

        comp, subcomp, agrup = _clasificar_inflacion(row, layout)

        fila: dict = {
            "generico": concepto,
            "ponderador": ponderador,
            "_grupo": grupo_actual,
            "inflacion componente": comp,
            "inflacion subcomponente": subcomp,
            "inflacion agrupacion": agrup,
        }

        if layout.col_encadenamiento:
            fila["encadenamiento"] = _celda(row, layout.col_encadenamiento)

        fila["canasta basica"] = "X" if _celda(row, layout.col_canasta_basica) == "X" else ""

        if layout.col_canasta_consumo_minimo:
            es_ccm = _celda(row, layout.col_canasta_consumo_minimo) == "X"
            fila["canasta consumo minimo"] = "X" if es_ccm else ""

        filas.append(fila)

    return pd.DataFrame(filas)


def _clasificar_inflacion(row: tuple, layout: LayoutXlsx) -> tuple[str, str, str]:
    """Busca la columna de agrupación marcada con X y devuelve su clasificación."""
    for col, (comp, subcomp, agrup) in layout.agrupaciones.items():
        if _celda(row, col) == "X":
            return comp, subcomp, agrup
    return "", "", ""


def _celda(row: tuple, col: int) -> object:
    """Valor de la celda en `col` (1-indexado, como en el xlsx)."""
    return row[col - 1].value


def _es_skip(texto: str) -> bool:
    """True si el texto es un renglón agregado a ignorar (total, suma, etc.)."""
    return texto.lower() in _SKIP


def _str(val: object) -> str:
    """Normaliza un valor de celda a string, vacío si es None."""
    if val is None:
        return ""
    return str(val).strip()
